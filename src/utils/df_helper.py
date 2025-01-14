import time

import pandas as pd
import openpyxl
import re

import unicodedata


def create_df_from_large_xlsx(
    file_name: str,
    sheet_name: str,
    start_row: int = 1,
    chunk_size: int = 1000,
    logging=None,
):
    try:
        inicio = time.time()
        workbook = openpyxl.load_workbook(file_name, data_only=True, read_only=True)
        worksheet = workbook.active if sheet_name == "" else workbook[sheet_name]
        fim = time.time()
        logging.info(f"Tempo de leitura do arquivo xlsx: {fim - inicio:.2f} segundos")
    except Exception as error:
        raise Exception(f"Erro ao tentar ler o arquivo {file_name}! {error}")

    inicio = time.time()
    try:
        df = read_excel_in_chunks(
            worksheet=worksheet,
            chunk_size=chunk_size,
            start_row=start_row,
        )
    except Exception as error:
        raise Exception(f"Erro ao tentar concatenar os DFs! {error}")

    fim = time.time()
    logging.info(f"Tempo de criacao dos DFs: {fim - inicio:.2f} segundos")

    return df


def read_excel_in_chunks(worksheet, chunk_size: int = 1000, start_row: int = 1):
    get_headers = None
    try:
        lista_linhas_worksheet = []
        df_list = []
        for row in worksheet.iter_rows(
            values_only=True, min_row=start_row, max_row=worksheet.max_row
        ):
            lista_linhas_worksheet.append(row)

            # Cada DF criado possui chunk_size observações.
            if len(lista_linhas_worksheet) >= chunk_size:
                if get_headers is None:
                    get_headers = lista_linhas_worksheet[0]
                    df = pd.DataFrame(lista_linhas_worksheet[1:], columns=get_headers)
                else:
                    df = pd.DataFrame(lista_linhas_worksheet, columns=get_headers)

                df_list.append(df)
                lista_linhas_worksheet.clear()
                del df

        if len(lista_linhas_worksheet) > 0:
            df = pd.DataFrame(lista_linhas_worksheet, columns=get_headers)
            df_list.append(df)
            del df

        df_concatenado = pd.concat(df_list, ignore_index=True)
    except Exception as error:
        raise Exception(f"Erro ao criar os DFs em chunks: {error}")

    return df_concatenado


def get_sheet_names(file_name):
    try:
        excel_file = pd.ExcelFile(file_name)
        sheet_names = excel_file.sheet_names
    except Exception as error:
        raise FileExistsError(
            f"Erro ao tentar carregar os nomes das planilhas do arquivo {file_name}! {error}"
        )

    return sheet_names


def validar_dado_numerico_como_string(string_data: str):
    try:
        if re.match(r"^\d{4}$", string_data) is not None:
            return True
        else:
            return False
    except Exception as error:
        raise Exception(
            f"Erro no metodo validar_dado_numerico_como_string ao tentar validar o ano {string_data}! {error}"
        )


def normalizar_string(text):
    try:
        return "".join(
            c
            for c in unicodedata.normalize("NFD", text)
            if unicodedata.category(c) != "Mn"
        ).lower()
    except Exception as error:
        raise Exception(f"Erro ao normalizar a string {text}! {error}")
