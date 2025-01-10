import pandas as pd
import openpyxl


def create_df_from_large_xlsx(file_name: str, start_row: int = 1, chunk_size: int = 1000):
    try:
        # Para criar um único DataFrame com todos os chunks:
        all_data = []
        for chunk in read_excel_in_chunks(
                file_name=file_name,
                chunk_size=chunk_size,
                start_row=start_row
        ):
            all_data.append(chunk)

        # une todos os dataframes da lista all_data
        df = pd.concat(all_data, ignore_index=True)
    except Exception as error:
        raise Exception(f"Erro ao tentar ler o arquivo! {error}")

    return df


def read_excel_in_chunks(file_name: str, chunk_size: int = 1000, start_row: int = 1):
    try:
        workbook = openpyxl.load_workbook(file_name, data_only=True)
        worksheet = workbook.active
    except Exception as error:
        raise Exception(f"Erro ao tentar ler o arquivo {file_name}! {error}")

    get_headers = None

    for i in range(start_row, worksheet.max_row, chunk_size):
        # Itera sobre as linhas da planilha, especificando o intervalo de linhas para cada chunk.
        rows = worksheet.iter_rows(min_row=i, max_row=min(i + chunk_size, worksheet.max_row))

        # Cria uma lista de listas, onde cada lista interna representa uma linha do chunk.
        data = [[cell.value for cell in row] for row in rows]

        if not get_headers:
            get_headers = data[0]
            data.pop(0)  # exclui o cabecalho do dataframe somente na primeira iteracao

        df_chunk = pd.DataFrame(data, columns=get_headers)
        """
        Retorna um gerador, permitindo processar os chunks de forma eficiente.
        Geradores permitem processar grandes quantidades de dados sem carregá-los todos de uma vez na memória.
        A palavra-chave yield transforma a função em um gerador. 
        Em cada iteração do loop, quando yield data é encontrado, o valor de data é retornado para o chamador da função.
        """
        yield df_chunk


def get_sheet_names(file_name):
    try:
        excel_file = pd.ExcelFile(file_name)
        sheet_names = excel_file.sheet_names
    except Exception as error:
        raise FileExistsError(f"Erro ao tentar carregar os nomes das planilhas do arquivo {file_name}! {error}")

    return sheet_names


"""
usecols=["IDADE","SEXO","CÓD.","SIGLA","LOCAL",
 "2000","2001","2002","2003","2004","2005",
 "2006","2007","2008","2009","2010","2011",
 "2012","2013","2014","2015","2016","2017",
 "2018","2019","2020","2021","2022","2023",
 "2024","2025","2026","2027","2028","2029",
 "2030","2031","2032","2033","2034","2035",
 "2036","2037","2038","2039","2040","2041",
 "2042","2043","2044","2045","2046","2047",
 "2048","2049","2050","2051","2052","2053",
 "2054","2055","2056","2057","2058","2059",
 "2060","2061","2062","2063","2064","2065",
 "2066","2067","2068","2069","2070"
 ],
dtype={
"IDADE":"str", "SEXO": "str", "CÓD.": "str",
"SIGLA": "str", "LOCAL": "str", "2000": "int32",
"2001": "int32", "2002": "int32", "2003": "int32",
"2004": "int32", "2005": "int32", "2006": "int32",
"2007": "int32", "2008": "int32", "2009": "int32",
"2010": "int32", "2011": "int32", "2012": "int32",
"2013": "int32", "2014": "int32", "2015": "int32",
"2016": "int32", "2017": "int32", "2018": "int32",
"2019": "int32", "2020": "int32", "2021": "int32",
"2022": "int32", "2023": "int32", "2024": "int32",
"2025": "int32", "2026": "int32", "2027": "int32",
"2028": "int32", "2029": "int32", "2030": "int32",
"2031": "int32", "2032": "int32", "2033": "int32",
"2034": "int32", "2035": "int32", "2036": "int32",
"2037": "int32", "2038": "int32", "2039": "int32",
"2040": "int32", "2041": "int32", "2042": "int32",
"2043": "int32", "2044": "int32", "2045": "int32",
"2046": "int32", "2047": "int32", "2048": "int32",
"2049": "int32", "2050": "int32", "2051": "int32",
"2052": "int32", "2053": "int32", "2054": "int32",
"2055": "int32", "2056": "int32", "2057": "int32",
"2058": "int32", "2059": "int32", "2060": "int32",
"2061": "int32", "2062": "int32", "2063": "int32",
"2064": "int32", "2065": "int32", "2066": "int32",
"2067": "int32", "2068": "int32", "2069": "int32",
"2070": "int32"
}
"""
